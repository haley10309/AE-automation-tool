"""
엑셀(xlsx) 파일 파싱 스크립트 — 두 단계 전략을 쓴다.

1) 먼저 openpyxl(순수 파이썬, Excel 앱/OS 무관)로 연다.
   대부분의 일반 파일은 이 단계에서 바로 성공하며 빠르고 가볍다.
2) openpyxl로 열리지 않으면(NASCA 등 DRM으로 보호된 파일은 정상적인
   zip/OOXML 구조가 아니라서 openpyxl이 예외를 던진다) xlwings(COM)로
   실제 설치된 Excel 애플리케이션을 띄워서 연다. 해당 PC에 NASCA
   플러그인이 설치돼 있으면 Excel이 파일을 여는 시점에 플러그인이
   자동으로 복호화를 처리해준다.

주의:
- 2단계(xlwings+COM)는 Windows 전용이다 (pywin32 필요). macOS/Linux에서는
  NASCA 같은 Windows 전용 DRM 파일 자체를 다룰 일이 없으므로, 1단계
  (openpyxl)만으로 충분하다. pywin32는 Windows에서 NASCA 파일을 열 때만
  필요하며, 그 외 환경에서는 설치하지 않아도 된다.
- 2단계는 DRM 복호화 특성상 openpyxl 대비 오픈 속도가 느릴 수 있다.
- 드물게 NASCA 플러그인이 최초 인증/권한 팝업을 띄우는 경우가 있는데,
  이 스크립트는 headless(visible=False)로 실행되므로 그런 팝업이 뜨면
  스크립트가 멈춘 것처럼 보일 수 있다. 이 경우 최초 1회는 사용자가
  화면에 보이는 Excel로 해당 파일을 직접 열어 인증을 완료해둬야 한다.
"""
import sys
import os
import json
import time
import traceback

MAX_ROWS = 3000  # 안전장치: 비정상적으로 큰 파일 방지
OPEN_RETRIES = 3       # NASCA 복호화 지연 대응 재시도 횟수
OPEN_RETRY_DELAY = 2   # 재시도 간 대기(초)

# 디버깅용: 환경변수 NASCA_DEBUG_VISIBLE=1로 실행하면 Excel 창을 화면에 띄워서
# NASCA 팝업이 실제로 뜨는지 눈으로 확인할 수 있다.
DEBUG_VISIBLE = os.environ.get('NASCA_DEBUG_VISIBLE') == '1'


def log(msg):
    """stderr로 즉시 flush 출력 — Node가 타임아웃으로 강제종료해도 그 전까지 찍힌 로그는 캡처됨"""
    print(msg, file=sys.stderr, flush=True)


def emit_error(msg, status_code=500):
    """항상 JSON을 stdout에 flush로 내보낸다 (Node가 stdout을 JSON.parse 하므로 필수)"""
    print(json.dumps({"errorMsg": msg, "statusCode": status_code}, ensure_ascii=False), flush=True)


def norm(v):
    if v is None:
        return ""
    return str(v).strip()


def open_book_with_retry(app, path):
    """NASCA 복호화가 느리게 끝나 첫 오픈이 실패하는 경우를 대비한 재시도"""
    last_err = None
    for attempt in range(1, OPEN_RETRIES + 1):
        try:
            return app.books.open(path)
        except Exception as e:
            last_err = e
            log(f"STEP:open 시도 {attempt}/{OPEN_RETRIES} 실패 - {e}")
            if attempt < OPEN_RETRIES:
                time.sleep(OPEN_RETRY_DELAY)
    raise last_err


def try_openpyxl(input_excel_path):
    """
    1단계: openpyxl로 파싱 시도 (Excel 앱 불필요, 모든 OS에서 동작, 빠름).
    NASCA 등 DRM으로 암호화된 파일은 정상적인 OOXML(zip) 구조가 아니므로
    이 함수가 예외를 던진다 — 그러면 호출부에서 xlwings 폴백으로 넘어간다.
    """
    import openpyxl

    wb = openpyxl.load_workbook(input_excel_path, data_only=True, read_only=False)
    try:
        sht = wb.worksheets[0]
        max_row = sht.max_row or 0
        max_col = sht.max_column or 0
        if max_row == 0 or max_col == 0:
            raise ValueError("시트에 데이터가 없습니다.")

        # 병합 범위를 먼저 수집 (읽는 동안 값을 채워 넣기 위함 — 실제 파일은 수정하지 않음)
        merge_ranges = list(sht.merged_cells.ranges)

        grid = [[None] * max_col for _ in range(max_row)]
        for row in sht.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                grid[cell.row - 1][cell.column - 1] = cell.value

        # 병합 해제 후 값 채우기와 동일한 효과: 병합 영역의 좌상단 값을 전체 셀에 복사
        for mrange in merge_ranges:
            min_col, min_row, max_col_m, max_row_m = mrange.bounds
            if min_row - 1 >= len(grid) or min_col - 1 >= len(grid[0]):
                continue
            top_val = grid[min_row - 1][min_col - 1]
            for r in range(min_row, min(max_row_m, max_row) + 1):
                for c in range(min_col, min(max_col_m, max_col) + 1):
                    grid[r - 1][c - 1] = top_val

        norm_grid = [[norm(v) for v in row] for row in grid]
        if len(norm_grid) > MAX_ROWS:
            norm_grid = norm_grid[:MAX_ROWS]

        return {
            "grid": norm_grid,
            "rowCount": len(norm_grid),
            "colCount": max((len(r) for r in norm_grid), default=0),
        }
    finally:
        wb.close()


def unmerge_and_fill(sheet):
    """
    병합된 셀 해제 후, 병합되어 있던 영역의 값을 모든 셀에 다시 채워 넣는다.
    (카테고리/키 컬럼이 여러 행에 걸쳐 병합되어 있는 케이스 대응.
     NASCA DRM과는 무관한, 카피덱 엑셀의 일반적인 서식 이슈 대응 로직)
    """
    used = sheet.api.UsedRange
    n_rows = used.Rows.Count
    n_cols = used.Columns.Count

    seen_addrs = set()
    merge_areas = []

    for r in range(1, n_rows + 1):
        for c in range(1, n_cols + 1):
            cell = used.Cells(r, c)
            if cell.MergeCells:
                area = cell.MergeArea
                addr = area.Address
                if addr not in seen_addrs:
                    seen_addrs.add(addr)
                    merge_areas.append(area)

    for area in merge_areas:
        val = area.Cells(1, 1).Value
        area.UnMerge()
        area.Value = val


def read_grid(sheet):
    used = sheet.used_range.value
    if not used:
        raise ValueError("시트에 데이터가 없습니다.")
    if not isinstance(used, list):
        # 셀 1개짜리 시트
        used = [[used]]
    if used and not isinstance(used[0], list):
        used = [used]

    if len(used) > MAX_ROWS:
        used = used[:MAX_ROWS]

    grid = [[norm(cell) for cell in row] for row in used]
    return grid


def run_with_xlwings(input_excel_path):
    """2단계: xlwings(COM, Windows 전용)로 실제 Excel 앱을 띄워서 연다.
    NASCA 등 DRM으로 보호된 파일이 openpyxl로 안 열릴 때만 여기로 온다."""
    import xlwings as xw  # 여기서 import — Windows가 아니거나 pywin32 미설치 시 에러를 명확히 전달하기 위해 지연 import

    app = xw.App(visible=DEBUG_VISIBLE, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    log(f"EXCEL_PID:{app.pid}")  # Node가 타임아웃 시 이 PID로 좀비 Excel 프로세스 강제종료
    wb = None

    try:
        log("STEP:opening file (NASCA 복호화 대기 구간)")
        wb = open_book_with_retry(app, input_excel_path)
        log("STEP:file opened, reading sheet")
        sht = wb.sheets[0]

        unmerge_and_fill(sht)
        log("STEP:merged cells unmerged, reading grid")
        grid = read_grid(sht)
        log("STEP:grid read complete")

        return {
            "grid": grid,
            "rowCount": len(grid),
            "colCount": max((len(r) for r in grid), default=0),
        }
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass
        try:
            app.quit()
        except Exception:
            pass


def run(input_excel_path):
    """
    1) openpyxl로 먼저 시도 (빠름, Excel 앱/OS 무관).
    2) 실패하면(주로 NASCA 등 DRM 파일) xlwings+COM(Windows 전용)으로 재시도.
    """
    try:
        log("STEP:openpyxl로 파싱 시도")
        result = try_openpyxl(input_excel_path)
        log("STEP:openpyxl 파싱 성공")
        print(json.dumps(result, ensure_ascii=False), flush=True)
        return
    except Exception as e:
        log(f"STEP:openpyxl 파싱 실패 - {type(e).__name__}: {e}")
        log("STEP:xlwings(Excel 자동화)로 재시도 — DRM 보호 파일일 가능성")

    # openpyxl로 못 읽은 경우에만 xlwings(Windows COM) 경로로 폴백
    result = run_with_xlwings(input_excel_path)
    print(json.dumps(result, ensure_ascii=False), flush=True)


def main():
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except Exception:
        pass

    if len(sys.argv) < 2:
        emit_error("Usage: python parse_merge_excel.py <input_excel_path>", 400)
        sys.exit(1)

    input_excel_path = sys.argv[1]

    if not os.path.exists(input_excel_path):
        emit_error(f"파일을 찾을 수 없습니다: {input_excel_path}", 404)
        sys.exit(1)

    try:
        run(input_excel_path)
    except ImportError as e:
        # 2단계(xlwings) import 실패 — 보통 이 파일이 NASCA 등 DRM으로 보호되어 있는데
        # 이 PC(비-Windows이거나 pywin32 미설치)에서는 xlwings COM 자동화를 쓸 수 없는 경우
        log(traceback.format_exc())
        emit_error(
            f"xlwings 모듈을 불러올 수 없습니다 ({e}). "
            f"이 파일은 openpyxl로 열리지 않아 xlwings(Excel 자동화)가 필요한데, "
            f"xlwings/pywin32는 Windows에서만 지원됩니다. "
            f"DRM(NASCA 등)으로 보호된 파일이라면 Windows PC에서 시도해주세요. "
            f"일반 파일인데 이 오류가 난다면 파일이 손상됐을 수 있습니다.",
            500,
        )
        sys.exit(1)
    except Exception as e:
        log(traceback.format_exc())
        emit_error(f"{type(e).__name__}: {e}", 500)
        sys.exit(1)


if __name__ == "__main__":
    main()